#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Aug  6 20:58:48 2023

@author: green-machine
"""

import os
import re
from pathlib import Path
from typing import Any, Union

from core.config import DATA_DIR
from openpyxl import load_workbook
from xlrd import open_workbook


def get_file_names(matchers):
    return [name for name in tuple(os.listdir()) if any(match in name for match in matchers)]


def get_file_names(path_src: str):
    return filter(
        lambda _: re.match('^dataset_che_(.*?)\.xls$', _), os.listdir(path_src)
    )


def get_xl_sheetnames(filepath: Union[str, Path]) -> list[str]:
    kwargs = {
        'filename': filepath,
        'readonly': True,
        'keep_links': False
    }
    return load_workbook(**kwargs).sheetnames


def get_xl_sheetnames(filepath: Union[str, Path]) -> list[str]:
    return open_workbook(filepath).sheet_names()


def archive_name_to_url(archive_name: str) -> str:
    """
    Parameters
    ----------
    archive_name : str
        DESCRIPTION.
    Returns
    -------
    str
        DESCRIPTION.
    """
    return f'https://www150.statcan.gc.ca/n1/tbl/csv/{archive_name}'


def dichotomize_series_ids(
    series_ids: dict[str, int],
    source_ids: tuple[int]
) -> tuple[dict[str, int]]:
    """
    Parameters
    ----------
    series_ids : dict[str, int]
        DESCRIPTION.
    source_ids : tuple[int]
        DESCRIPTION.
    Returns
    -------
    tuple[dict[str, int]]
        DESCRIPTION.f
    """
    return (
        {
            key: value for key, value in series_ids.items() if not value in source_ids
        },
        {
            key: value for key, value in series_ids.items() if value in source_ids
        }
    )


def get_pre_kwargs(file_name: str) -> dict[str, Any]:
    """
    Returns `kwargs` for `pd.read_csv()` for Usual Cases

    Parameters
    ----------
    file_name : str
        DESCRIPTION.

    Returns
    -------
    dict[str, Any]
        DESCRIPTION.

    """
    return {
        'filepath_or_buffer': DATA_DIR.joinpath(file_name),
        'index_col': 0,
    }


def get_kwargs_usa_frb_gvp_sa() -> dict[str, Any]:
    """'T50030: Final products and nonindustrial supplies--gross value'"""
    FILE_NAME = 'dataset_usa_frb_gvp_sa.txt'
    return {
        'filepath_or_buffer': DATA_DIR.joinpath(FILE_NAME),
        'sep': '\s+',
        'header': None,
        'skiprows': 1,
        'nrows': 29,
        'index_col': 1
    }


def get_kwargs_usa_bls_ap_pc() -> dict[str, Any]:

    FILE_NAME = 'usa_science_data.zip/pc.df.0.Current' or 'usa_science_data.zip/ap.df.0.Current'

    return {
        'filepath_or_buffer': DATA_DIR.joinpath(FILE_NAME),
        'sep': '\t',
        'index_col': range(4),
        'usecols': 1,
    }
