import os
import re
from pathlib import PosixPath
from typing import Union

from openpyxl import load_workbook
from xlrd import open_workbook


def get_file_names(matchers):
    return [name for name in tuple(os.listdir()) if any(match in name for match in matchers)]


def get_file_names(path_src: str):
    return filter(
        lambda _: re.match('^dataset_che_(.*?)\.xls$', _), os.listdir(path_src)
    )


def get_xl_sheetnames(filepath: Union[str, PosixPath]) -> list[str]:
    kwargs = {
        'filename': filepath,
        'readonly': True,
        'keep_links': False
    }
    return load_workbook(**kwargs).sheetnames


def get_xl_sheetnames(filepath: Union[str, PosixPath]) -> list[str]:
    return open_workbook(filepath).sheet_names()


def archive_name_to_url(archive_name: str) -> str:
    """
    Parameters
    ----------
    archive_name : str
        DESCRIPTION.
    Returns
    -------
    str
        DESCRIPTION.
    """
    return f'https://www150.statcan.gc.ca/n1/tbl/csv/{archive_name}'


def dichotomize_series_ids(
    series_ids: dict[str, int],
    source_ids: tuple[int]
) -> tuple[dict[str, int]]:
    """
    Parameters
    ----------
    series_ids : dict[str, int]
        DESCRIPTION.
    source_ids : tuple[int]
        DESCRIPTION.
    Returns
    -------
    tuple[dict[str, int]]
        DESCRIPTION.
    """
    return (
        {
            key: value for key, value in series_ids.items() if not value in source_ids
        },
        {
            key: value for key, value in series_ids.items() if value in source_ids
        }
    )
