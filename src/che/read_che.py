# -*- coding: utf-8 -*-
"""
Created on Tue Apr 28 00:37:30 2020

@author: Alexander Mikhailov
"""

import os
import re
from pathlib import Path, PosixPath
from typing import Union

import pandas as pd
from openpyxl import load_workbook
from xlrd import open_workbook


def get_xl_sheetnames(filepath: Union[str, PosixPath]) -> list[str]:
    kwargs = {
        'filename': filepath,
        'read_only': True,
        'keep_links': False
    }
    return load_workbook(**kwargs).sheetnames


def get_xl_sheetnames(filepath: Union[str, PosixPath]) -> list[str]:
    return open_workbook(filepath).sheet_names()


def main(path_src: str = '/media/green-machine/KINGSTON'):
    file_names = filter(
        lambda _: re.match('^dataset_che_(.*?)\.xls$', _), os.listdir(path_src)
    )

    for file_name in file_names:
        print(file_name)
        for sheet_name in get_xl_sheetnames(Path(path_src).joinpath(file_name)):
            print(pd.read_excel(Path(path_src).joinpath(file_name), sheet_name))


if __name__ == '__main__':
    main()
